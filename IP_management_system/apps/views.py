from django.shortcuts import render, redirect
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from apps import models

from datetime import datetime
import os
import time
import re


# 网络地址对应处理

def ip_list(request):
    all_ip = models.AppIpTest.objects.all()
    paginator = Paginator(all_ip, 20)
    current_page = int(request.GET.get("page", 1))
    # 页对象
    page = paginator.page(current_page)

    # 构建page_range
    max_page_count = 20
    max_page_count_half = int(max_page_count / 2)
    # 判断页数是否大于max_page_count
    if paginator.num_pages >= max_page_count:
        # 得出start位置
        if current_page <= max_page_count_half:
            page_start = 1
            page_end = max_page_count + 1
        else:
            if current_page + max_page_count_half + 1 > paginator.num_pages:
                page_start = paginator.num_pages - max_page_count
                page_end = paginator.num_pages + 1
            else:
                page_start = current_page - max_page_count_half
                page_end = current_page + max_page_count_half + 1
        page_range = range(page_start, page_end)
    else:
        page_range = paginator.page_range
    # return render(request, 'ip_list.html', {'ips': all_ip,'ips_no':ips_no})
    return render(request, 'ip_list.html', locals())


def add_ip(request):
    error_msg = ''
    ret = models.AppIpTest.objects.all()
    if request.method == 'POST':
        new_address = request.POST.get('ip_address')
        if new_address:
            new_name = request.POST.get('ip_name')
            new_mac = request.POST.get('ip_mac')
            new_user = request.POST.get('ip_user')
            new_status = request.POST.get('ip_status')
            new_description = request.POST.get('ip_change')
            models.AppIpTest.objects.create(address=new_address, name=new_name, mac=new_mac, user=new_user, status=new_status, description=new_description)
            return redirect('/ip_list/')
        else:
            error_msg = '地址不能为空'
    return render(request, 'add_ip.html', {'ip_list': ret, 'error_msg': error_msg})


def delete_ip(request, del_id):
    models.AppIpTest.objects.get(id=del_id).delete()
    return redirect('/ip_list/')

#
def edit_ip(request, edit_id):
    error_msg = ''
    if request.method == 'POST':
        new_address = request.POST.get('ip_address')
        if new_address:
            new_user = request.POST.get('ip_user')
            new_name = request.POST.get('ip_name')
            new_mac = request.POST.get('ip_mac')
            new_status = request.POST.get('ip_status')
            new_description = request.POST.get('ip_description')
            edit_ip_obj = models.AppIpTest.objects.get(id=edit_id)
            edit_ip_obj.address = new_address
            edit_ip_obj.name = new_name
            edit_ip_obj.mac = new_mac
            edit_ip_obj.user = new_user
            edit_ip_obj.status = new_status
            edit_ip_obj.description = new_description
            edit_ip_obj.save()
            return redirect('/ip_list/')
        else:
            error_msg = '书名不能为空'
    edit_ip_obj = models.AppIpTest.objects.get(id=edit_id)
    return render(request, 'edit_ip.html',
                  {'ip_list': ip_list, 'edit_ip_obj': edit_ip_obj, 'error_msg': error_msg})


def search(request):
    if request.method == 'POST':
        search_address = request.POST.get('search_address', '')
        search_name = request.POST.get('search_name', '')
        search_user = request.POST.get('search_user', '')
        search_status = request.POST.get('search_status', '')
        filter_ip = models.AppIpTest.objects.filter(address__icontains=search_address, name__icontains=search_name,
                                          user__icontains=search_user, status__icontains=search_status)
        print('1111')
        print(filter_ip)
        print('1111')
        paginator = Paginator(filter_ip, 1000)
        current_page = int(request.GET.get("page", 1))
        # 页对象
        page = paginator.page(current_page)

        # 构建page_range
        max_page_count = 1000
        max_page_count_half = int(max_page_count / 2)
        # 判断页数是否大于max_page_count
        if paginator.num_pages >= max_page_count:
            # 得出start位置
            if current_page <= max_page_count_half:
                page_start = 1
                page_end = max_page_count + 1
            else:
                if current_page + max_page_count_half + 1 > paginator.num_pages:
                    page_start = paginator.num_pages - max_page_count
                    page_end = paginator.num_pages + 1
                else:
                    page_start = current_page - max_page_count_half
                    page_end = current_page + max_page_count_half + 1
            page_range = range(page_start, page_end)
        else:
            page_range = paginator.page_range
        print('2222')
        print(locals())
        print('2222')
        return render(request, 'ip_list.html', locals())


def download(request):
    import datetime
    from django.http import HttpResponse
    from openpyxl import Workbook
    from io import BytesIO
    from django.utils.http import urlquote
    wb = Workbook()
    wb.encoding = 'utf-8'
    sheet1 = wb.active
    row_one = ["id", "IP地址", "主机名", "MAC地址", "用户", "状态", "详细描述"]
    for i in range(1, len(row_one)+1):
        sheet1.cell(row=1, column=i).value=row_one[i-1]
    all_ip = models.AppIpTest.objects.all()
    for ip in all_ip:
        max_row = sheet1.max_row + 1
        ip_info = [ip.id, ip.address, ip.name, ip.mac, ip.user, ip.status, ip.description]
        for x in range(1, len(ip_info)+1):
            sheet1.cell(row=max_row, column=x).value=ip_info[x-1]
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    response = HttpResponse(output.getvalue(), content_type='application/vnd.ms-excel')
    timestamp = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    file_name = "IP_list%s.xls" % timestamp
    file_name = urlquote(file_name)
    response['Content-Disposition'] = 'attachment; filename=%s' % file_name
    response.write(output.getvalue())
    return response
		
		
		
def index(request):
    return render(request, 'index.html')
